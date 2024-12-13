import mysql.connector
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# MySQL connection details
db_config = {
    "host": "localhost",
    "database": "ja_db",
    "user": "root",
    "password": "2044",
}

# Establish a database connection
conn = mysql.connector.connect(**db_config)

# Query to fetch transaction data
query = """
SELECT TransactionDate, SupplierId, `Value`
FROM transactions
WHERE TransactionType = 'G'
"""

# Load data into a pandas DataFrame
df = pd.read_sql(query, conn)

# Close the database connection
conn.close()

# Ensure TransactionDate is datetime
df["TransactionDate"] = pd.to_datetime(df["TransactionDate"])

# Add a column for year-month
df["YearMonth"] = df["TransactionDate"].dt.to_period("M")

# Sort by SupplierId, YearMonth, and TransactionDate
df = df.sort_values(by=["SupplierId", "YearMonth", "TransactionDate"])

# Get the first transaction of each month for each supplier
opening_balances = df.groupby(["SupplierId", "YearMonth"]).first().reset_index()

# Keep only the necessary columns
opening_balances = opening_balances[["YearMonth", "SupplierId", "Value"]]

# Convert YearMonth (Period) to string
opening_balances["YearMonth"] = opening_balances["YearMonth"].astype(str)

# Save the result to an Excel file
file_path = "opening_balances.xlsx"

# Create a workbook and add the DataFrame
wb = Workbook()
ws = wb.active
ws.title = "Opening Balances"

# Write DataFrame to the sheet
for r_idx, row in enumerate(
    dataframe_to_rows(opening_balances, index=False, header=True), start=1
):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Add a bar chart
chart = BarChart()
chart.title = "Opening Balances"
chart.x_axis.title = "Supplier and Month"
chart.y_axis.title = "Value"

# Define data for the chart
data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row, max_col=3)  # Values
categories = Reference(
    ws, min_col=1, min_row=2, max_row=ws.max_row
)  # YearMonth-Supplier pairs
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Place the chart on the sheet
ws.add_chart(chart, "E5")

# Save the Excel file
wb.save(file_path)

print(f"Excel file with chart saved as {file_path}")
